import os
import openpyxl
from io import BytesIO
from celery import shared_task
from django.core.files.base import ContentFile
from django.utils import timezone
from users.models import User
from bookings.models import Booking
from special_events.models import SpecialEvent, SpecialEventRegistration
from admin_panel.models import BulkUploadTask
from decimal import Decimal, InvalidOperation


def _cell_str(val):
    """Return a stripped string or empty string for a cell value."""
    if val is None:
        return ''
    return str(val).strip()


def _make_header_map(headers):
    """Return {lower_stripped_header: 0-based-index} for non-None headers."""
    return {str(h).strip().lower(): idx for idx, h in enumerate(headers) if h}


def _find_col(header_map, *keys):
    """Find the first column index matching any of the given substrings."""
    for key in keys:
        for k, v in header_map.items():
            if key in k:
                return v
    return None


def _parse_datetime(val, location_id=None):
    """
    Parse a cell value into an aware UTC datetime.

    The admin enters times in the golf club's LOCAL timezone (e.g. 09:00 Halifax).
    We must convert that local time to UTC before storing it, exactly as the rest
    of the booking system does via local_to_utc() in timezone_utils.
    """
    from datetime import datetime
    from golf_project.timezone_utils import local_to_utc

    if val is None:
        raise ValueError("Datetime cell is empty")

    # openpyxl may return a datetime object directly from the cell
    if hasattr(val, 'hour'):  # it's a datetime-like object
        naive_local = val.replace(tzinfo=None)  # strip any tz — treat as local
        return local_to_utc(naive_local, location_id)

    # String fallback
    s = str(val).strip()
    for fmt in ('%Y-%m-%d %H:%M', '%Y-%m-%d %H:%M:%S', '%Y-%m-%dT%H:%M', '%Y-%m-%dT%H:%M:%S'):
        try:
            naive_local = datetime.strptime(s, fmt)
            return local_to_utc(naive_local, location_id)
        except ValueError:
            continue
    raise ValueError(f"Cannot parse datetime '{s}'. Expected format: YYYY-MM-DD HH:MM (in the club's local time).")


def _parse_date(val):
    """Parse a cell value into a date object."""
    from datetime import date
    if val is None:
        raise ValueError("Date cell is empty")
    if hasattr(val, 'date') and callable(val.date):
        return val.date()
    if isinstance(val, date):
        return val
    s = str(val).strip()
    for fmt in ('%Y-%m-%d', '%d/%m/%Y', '%m/%d/%Y'):
        try:
            from datetime import datetime as dt
            return dt.strptime(s, fmt).date()
        except ValueError:
            continue
    raise ValueError(f"Cannot parse date '{s}'. Use YYYY-MM-DD format.")


@shared_task
def process_bulk_upload_task(task_id, file_path):
    try:
        task = BulkUploadTask.objects.get(id=task_id)
        task.status = 'processing'
        task.save()

        wb = openpyxl.load_workbook(filename=file_path, data_only=True)

        expected_sheets = ['Customers', 'Packages', 'Bookings', 'Events']
        total_rows = 0
        for sheet_name in expected_sheets:
            if sheet_name in wb.sheetnames:
                sheet = wb[sheet_name]
                # rows 1 = header, 2 = mandatory label, 3 = example → data from row 4
                data_rows = max(0, sheet.max_row - 3)
                total_rows += data_rows

        task.total_rows = total_rows
        task.save()

        error_wb = openpyxl.Workbook()
        error_wb.remove(error_wb.active)
        has_errors = False
        processed_count = 0

        # ──────────────────────────────────────────────────────────────────────
        # 1. CUSTOMERS
        # ──────────────────────────────────────────────────────────────────────
        if 'Customers' in wb.sheetnames:
            sheet = wb['Customers']
            err_sheet = error_wb.create_sheet('Customers')
            raw_headers = [cell.value for cell in sheet[1]]
            err_sheet.append(raw_headers + ['Error Message'])
            header_map = _make_header_map(raw_headers)

            name_idx  = _find_col(header_map, 'name', 'full name')
            phone_idx = _find_col(header_map, 'phone')
            email_idx = _find_col(header_map, 'email')
            dob_idx   = _find_col(header_map, 'date of birth', 'dob', 'birth')

            # Skip rows 2 (mandatory labels) and 3 (example) → start at row 4
            for row in sheet.iter_rows(min_row=4, values_only=True):
                if not any(row):
                    continue
                try:
                    if name_idx is None:
                        raise ValueError("'Full Name' column not found in Customers sheet")
                    if phone_idx is None:
                        raise ValueError("'Phone' column not found in Customers sheet")

                    name  = _cell_str(row[name_idx])
                    phone = _cell_str(row[phone_idx])
                    email = _cell_str(row[email_idx]) if email_idx is not None else ''

                    if not name:
                        raise ValueError("Full Name is mandatory")
                    if not phone:
                        raise ValueError("Phone is mandatory")

                    parts = name.split(' ', 1)
                    first_name = parts[0]
                    last_name  = parts[1] if len(parts) > 1 else ''

                    dob = None
                    if dob_idx is not None and row[dob_idx]:
                        try:
                            dob = _parse_date(row[dob_idx])
                        except ValueError:
                            pass  # DOB is optional — ignore parse failures

                    user, created = User.objects.get_or_create(
                        phone=phone,
                        defaults={
                            'username': phone,
                            'first_name': first_name,
                            'last_name': last_name,
                            'email': email,
                            'role': 'client',
                            'email_verified': True,
                            'phone_verified': True,
                            'ghl_location_id': task.location_id,
                            **(({'date_of_birth': dob}) if dob else {}),
                        }
                    )
                    if not created:
                        # Update name/email if they were empty
                        changed = False
                        if not user.first_name and first_name:
                            user.first_name = first_name
                            changed = True
                        if not user.last_name and last_name:
                            user.last_name = last_name
                            changed = True
                        if not user.email and email:
                            user.email = email
                            changed = True
                        if changed:
                            user.save(update_fields=['first_name', 'last_name', 'email'])

                except Exception as e:
                    err_sheet.append(list(row) + [str(e)])
                    has_errors = True

                processed_count += 1
                if processed_count % 10 == 0:
                    task.processed_rows = processed_count
                    task.save()

        # ──────────────────────────────────────────────────────────────────────
        # 2. PACKAGES
        # ──────────────────────────────────────────────────────────────────────
        if 'Packages' in wb.sheetnames:
            from coaching.models import CoachingPackage, CoachingPackagePurchase, SimulatorPackage, SimulatorPackagePurchase
            sheet = wb['Packages']
            err_sheet = error_wb.create_sheet('Packages')
            raw_headers = [cell.value for cell in sheet[1]]
            err_sheet.append(raw_headers + ['Error Message'])
            header_map = _make_header_map(raw_headers)

            phone_idx    = _find_col(header_map, 'phone', 'customer phone')
            pkg_idx      = _find_col(header_map, 'package name')
            type_idx     = _find_col(header_map, 'package type', 'type')
            sessions_idx = _find_col(header_map, 'sessions remaining', 'sessions')
            hours_idx    = _find_col(header_map, 'hours remaining', 'hours')
            notes_idx    = _find_col(header_map, 'notes', 'label', 'purchase label')

            for row in sheet.iter_rows(min_row=4, values_only=True):
                if not any(row):
                    continue
                try:
                    if phone_idx is None:
                        raise ValueError("'Customer Phone' column not found")
                    if pkg_idx is None:
                        raise ValueError("'Package Name' column not found")
                    if type_idx is None:
                        raise ValueError("'Package Type' column not found")

                    phone    = _cell_str(row[phone_idx])
                    pkg_name = _cell_str(row[pkg_idx])
                    pkg_type = _cell_str(row[type_idx]).lower()
                    notes    = _cell_str(row[notes_idx]) if notes_idx is not None and row[notes_idx] else ''

                    if not phone:
                        raise ValueError("Customer Phone is mandatory")
                    if not pkg_name:
                        raise ValueError("Package Name is mandatory")
                    if pkg_type not in ('coaching', 'simulator'):
                        raise ValueError(f"Package Type must be 'coaching' or 'simulator', got '{pkg_type}'")

                    user = User.objects.filter(phone=phone).first()
                    if not user:
                        raise ValueError(f"No customer found with phone '{phone}'. Add them in the Customers tab first.")

                    if pkg_type == 'coaching':
                        package = CoachingPackage.objects.filter(
                            title__iexact=pkg_name,
                            location_id=task.location_id
                        ).first() or CoachingPackage.objects.filter(title__iexact=pkg_name).first()
                        if not package:
                            raise ValueError(f"Coaching package '{pkg_name}' not found in the system")

                        # Sessions remaining: use provided value or default to full package
                        sessions_val = row[sessions_idx] if sessions_idx is not None else None
                        if sessions_val is not None and str(sessions_val).strip():
                            try:
                                sessions_remaining = int(str(sessions_val).strip())
                            except (ValueError, TypeError):
                                raise ValueError(f"Sessions Remaining must be a whole number, got '{sessions_val}'")
                        else:
                            sessions_remaining = package.session_count

                        sessions_total = max(package.session_count, sessions_remaining)

                        CoachingPackagePurchase.objects.create(
                            client=user,
                            package=package,
                            sessions_total=sessions_total,
                            sessions_remaining=sessions_remaining,
                            simulator_hours_total=package.simulator_hours,
                            simulator_hours_remaining=package.simulator_hours,
                            category_hours_total=package.category_hours,
                            category_hours_remaining=package.category_hours,
                            purchase_name=notes or f"Imported — {pkg_name}",
                            notes=notes,
                            package_status='active' if sessions_remaining > 0 else 'completed',
                        )

                    else:  # simulator
                        package = SimulatorPackage.objects.filter(
                            title__iexact=pkg_name,
                            location_id=task.location_id
                        ).first() or SimulatorPackage.objects.filter(title__iexact=pkg_name).first()
                        if not package:
                            raise ValueError(f"Simulator package '{pkg_name}' not found in the system")

                        hours_val = row[hours_idx] if hours_idx is not None else None
                        if hours_val is not None and str(hours_val).strip():
                            try:
                                hours_remaining = Decimal(str(hours_val).strip())
                            except InvalidOperation:
                                raise ValueError(f"Hours Remaining must be a number, got '{hours_val}'")
                        else:
                            hours_remaining = package.hours

                        hours_total = max(package.hours, hours_remaining)

                        SimulatorPackagePurchase.objects.create(
                            client=user,
                            package=package,
                            hours_total=hours_total,
                            hours_remaining=hours_remaining,
                            purchase_name=notes or f"Imported — {pkg_name}",
                            notes=notes,
                            package_status='active' if hours_remaining > 0 else 'completed',
                        )

                except Exception as e:
                    err_sheet.append(list(row) + [str(e)])
                    has_errors = True

                processed_count += 1
                if processed_count % 10 == 0:
                    task.processed_rows = processed_count
                    task.save()

        # ──────────────────────────────────────────────────────────────────────
        # 3. BOOKINGS
        # ──────────────────────────────────────────────────────────────────────
        if 'Bookings' in wb.sheetnames:
            from categories.models import ServiceCategory
            sheet = wb['Bookings']
            err_sheet = error_wb.create_sheet('Bookings')
            raw_headers = [cell.value for cell in sheet[1]]
            err_sheet.append(raw_headers + ['Error Message'])
            header_map = _make_header_map(raw_headers)

            phone_idx    = _find_col(header_map, 'phone', 'customer phone')
            type_idx     = _find_col(header_map, 'booking type', 'type')
            cat_idx      = _find_col(header_map, 'category name', 'category')
            start_idx    = _find_col(header_map, 'start time', 'start')
            end_idx      = _find_col(header_map, 'end time', 'end')
            status_idx   = _find_col(header_map, 'status')
            price_idx    = _find_col(header_map, 'total price', 'price')

            VALID_BOOKING_TYPES   = {'simulator', 'coaching', 'category'}
            VALID_BOOKING_STATUSES = {'completed', 'confirmed', 'cancelled', 'no_show'}

            for row in sheet.iter_rows(min_row=4, values_only=True):
                if not any(row):
                    continue
                try:
                    if phone_idx is None:
                        raise ValueError("'Customer Phone' column not found")
                    if type_idx is None:
                        raise ValueError("'Booking Type' column not found")
                    if start_idx is None:
                        raise ValueError("'Start Time' column not found")
                    if end_idx is None:
                        raise ValueError("'End Time' column not found")

                    phone    = _cell_str(row[phone_idx])
                    b_type   = _cell_str(row[type_idx]).lower()
                    cat_name = _cell_str(row[cat_idx]) if cat_idx is not None else ''
                    b_status = _cell_str(row[status_idx]).lower() if status_idx is not None and row[status_idx] else 'completed'

                    if not phone:
                        raise ValueError("Customer Phone is mandatory")
                    if not b_type:
                        raise ValueError("Booking Type is mandatory")
                    if b_type not in VALID_BOOKING_TYPES:
                        raise ValueError(f"Booking Type must be one of {sorted(VALID_BOOKING_TYPES)}, got '{b_type}'")
                    if b_status not in VALID_BOOKING_STATUSES:
                        b_status = 'completed'

                    # Parse as LOCAL club time and convert to UTC (matching system convention)
                    start_time = _parse_datetime(row[start_idx], location_id=task.location_id)
                    end_time   = _parse_datetime(row[end_idx],   location_id=task.location_id)

                    if end_time <= start_time:
                        raise ValueError("End Time must be after Start Time")

                    try:
                        total_price = Decimal(str(row[price_idx]).strip()) if price_idx is not None and row[price_idx] else Decimal('0.00')
                    except InvalidOperation:
                        total_price = Decimal('0.00')

                    user = User.objects.filter(phone=phone).first()
                    if not user:
                        raise ValueError(f"No customer found with phone '{phone}'")

                    # Resolve service_category for 'category' type
                    service_category = None
                    if b_type == 'category':
                        if not cat_name:
                            raise ValueError("Category Name is required when Booking Type is 'category'")
                        service_category = ServiceCategory.objects.filter(
                            name__iexact=cat_name,
                            location_id=task.location_id
                        ).first() or ServiceCategory.objects.filter(name__iexact=cat_name).first()
                        if not service_category:
                            raise ValueError(f"Service category '{cat_name}' not found in the system")

                    # Map 'category' to 'simulator' for DB storage (legacy field)
                    db_booking_type = b_type if b_type in ('simulator', 'coaching') else 'simulator'

                    Booking.objects.create(
                        client=user,
                        location_id=task.location_id,
                        booking_type=db_booking_type,
                        status=b_status,
                        start_time=start_time,
                        end_time=end_time,
                        total_price=total_price,
                        service_category=service_category,
                    )

                except Exception as e:
                    err_sheet.append(list(row) + [str(e)])
                    has_errors = True

                processed_count += 1
                if processed_count % 10 == 0:
                    task.processed_rows = processed_count
                    task.save()

        # ──────────────────────────────────────────────────────────────────────
        # 4. EVENTS
        # ──────────────────────────────────────────────────────────────────────
        if 'Events' in wb.sheetnames:
            sheet = wb['Events']
            err_sheet = error_wb.create_sheet('Events')
            raw_headers = [cell.value for cell in sheet[1]]
            err_sheet.append(raw_headers + ['Error Message'])
            header_map = _make_header_map(raw_headers)

            phone_idx  = _find_col(header_map, 'phone', 'customer phone')
            event_idx  = _find_col(header_map, 'event name', 'event')
            date_idx   = _find_col(header_map, 'occurrence date', 'date')
            status_idx = _find_col(header_map, 'registration status', 'status')

            VALID_REG_STATUSES = {'showed_up', 'registered', 'cancelled'}

            for row in sheet.iter_rows(min_row=4, values_only=True):
                if not any(row):
                    continue
                try:
                    if phone_idx is None:
                        raise ValueError("'Customer Phone' column not found")
                    if event_idx is None:
                        raise ValueError("'Event Name' column not found")
                    if date_idx is None:
                        raise ValueError("'Occurrence Date' column not found")

                    phone      = _cell_str(row[phone_idx])
                    event_name = _cell_str(row[event_idx])
                    reg_status = _cell_str(row[status_idx]).lower() if status_idx is not None and row[status_idx] else 'showed_up'

                    if not phone:
                        raise ValueError("Customer Phone is mandatory")
                    if not event_name:
                        raise ValueError("Event Name is mandatory")

                    if reg_status not in VALID_REG_STATUSES:
                        reg_status = 'showed_up'

                    occ_date = _parse_date(row[date_idx])

                    user = User.objects.filter(phone=phone).first()
                    if not user:
                        raise ValueError(f"No customer found with phone '{phone}'")

                    event = SpecialEvent.objects.filter(
                        title__iexact=event_name,
                        location_id=task.location_id
                    ).first() or SpecialEvent.objects.filter(title__iexact=event_name).first()
                    if not event:
                        raise ValueError(f"Special event '{event_name}' not found in the system")

                    SpecialEventRegistration.objects.get_or_create(
                        event=event,
                        user=user,
                        occurrence_date=occ_date,
                        defaults={'status': reg_status}
                    )

                except Exception as e:
                    err_sheet.append(list(row) + [str(e)])
                    has_errors = True

                processed_count += 1
                if processed_count % 10 == 0:
                    task.processed_rows = processed_count
                    task.save()

        # ──────────────────────────────────────────────────────────────────────
        # Finalize
        # ──────────────────────────────────────────────────────────────────────
        if has_errors:
            out = BytesIO()
            error_wb.save(out)
            task.error_file.save(f"errors_{task_id}.xlsx", ContentFile(out.getvalue()))

        task.status = 'completed'
        task.processed_rows = total_rows
        task.save()

        if os.path.exists(file_path):
            os.remove(file_path)

    except Exception as e:
        try:
            task = BulkUploadTask.objects.get(id=task_id)
            task.status = 'failed'
            error_wb2 = openpyxl.Workbook()
            ws = error_wb2.active
            ws.append(["Critical Error"])
            ws.append([str(e)])
            out = BytesIO()
            error_wb2.save(out)
            task.error_file.save(f"critical_error_{task_id}.xlsx", ContentFile(out.getvalue()))
            task.save()
        except Exception:
            pass

        if os.path.exists(file_path):
            os.remove(file_path)
