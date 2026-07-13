import os
import django
import sys
from datetime import datetime, timedelta

# Setup Django environment
sys.path.append(os.path.dirname(os.path.abspath(__file__)))
os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'golf_project.settings')
django.setup()

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from coaching.models import CoachingPackage, SimulatorPackage
from categories.models import ServiceCategory
from special_events.models import SpecialEvent
from django.utils import timezone

LOCATION_ID = "test_location_123"

def setup_dummy_data():
    """Create dummy packages, categories, and events in the DB so the success upload actually works."""
    print("Setting up dummy database records for testing...")
    
    # 1. Coaching Package
    cp, _ = CoachingPackage.objects.get_or_create(
        title="10-Session Pro Coaching",
        defaults={
            'description': "10 hours of pro coaching",
            'price': 1000.00,
            'session_count': 10,
            'location_id': LOCATION_ID
        }
    )
    
    # 2. Simulator Package
    sp, _ = SimulatorPackage.objects.get_or_create(
        title="20-Hour Winter Pass",
        defaults={
            'description': "20 hours of simulator time",
            'price': 800.00,
            'hours': 20,
            'location_id': LOCATION_ID
        }
    )
    
    # 3. Service Category
    sc, _ = ServiceCategory.objects.get_or_create(
        name="Table Tennis",
        defaults={
            'slug': 'table-tennis',
            'customer_label': 'Table Tennis',
            'location_id': LOCATION_ID,
            'legacy_booking_type': 'simulator'
        }
    )
    
    # 4. Special Event
    today = timezone.now().date()
    ev, _ = SpecialEvent.objects.get_or_create(
        title="Saturday Morning Clinic",
        defaults={
            'description': "Weekly group clinic",
            'date': today,
            'start_time': '09:00:00',
            'end_time': '10:00:00',
            'max_capacity': 20,
            'location_id': LOCATION_ID
        }
    )
    return cp, sp, sc, ev

def create_excel(filename, is_success=True):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    
    # --- Style Helpers ---
    header_fill = PatternFill("solid", fgColor="1F4E79")
    mandatory_fill = PatternFill("solid", fgColor="FCE4D6")
    optional_fill = PatternFill("solid", fgColor="E2EFDA")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    mandatory_font = Font(bold=True, color="843C0C", size=10)
    optional_font = Font(bold=True, color="375623", size=10)
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border_side = Side(style="thin", color="BFBFBF")
    thin_border = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thin_border_side)

    def style_header(ws, columns):
        ws.row_dimensions[1].height = 36
        for col_idx, (header, is_mandatory, width) in enumerate(columns, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_align
            cell.border = thin_border
            ws.column_dimensions[get_column_letter(col_idx)].width = width
        ws.row_dimensions[2].height = 20
        for col_idx, (_, is_mandatory, _) in enumerate(columns, 1):
            cell = ws.cell(row=2, column=col_idx, value="MANDATORY" if is_mandatory else "optional")
            cell.fill = mandatory_fill if is_mandatory else optional_fill
            cell.font = mandatory_font if is_mandatory else optional_font
            cell.alignment = center_align
            cell.border = thin_border

    def add_rows(ws, rows_data):
        for row_idx, row_values in enumerate(rows_data, 4):  # Start at row 4
            for col_idx, val in enumerate(row_values, 1):
                ws.cell(row=row_idx, column=col_idx, value=val)

    # -------------------
    # Customers
    # -------------------
    ws_cust = wb.create_sheet("Customers")
    cust_cols = [("Full Name", True, 22), ("Phone", True, 18), ("Email", False, 28), ("Date of Birth", False, 18)]
    style_header(ws_cust, cust_cols)
    
    cust_rows = []
    for i in range(1, 11):
        if is_success:
            cust_rows.append([f"Test User {i}", f"+155500000{i:02d}", f"test{i}@example.com", f"1990-01-{i:02d}"])
        else:
            # Add some intentional errors in the failure file
            if i == 1:
                cust_rows.append(["", "+15550000001", "missingname@example.com", ""]) # Error: Missing Name
            elif i == 2:
                cust_rows.append(["No Phone User", "", "nophone@example.com", ""]) # Error: Missing Phone
            elif i == 3:
                cust_rows.append(["Invalid Date User", "+15550000003", "", "Not a date"]) # Warning/Ignored: Invalid date format
            else:
                cust_rows.append([f"Test User {i}", f"+155500000{i:02d}", f"test{i}@example.com", f"1990-01-{i:02d}"])

    add_rows(ws_cust, cust_rows)

    # -------------------
    # Packages
    # -------------------
    ws_pkg = wb.create_sheet("Packages")
    pkg_cols = [("Customer Phone", True, 18), ("Package Name", True, 30), ("Package Type", True, 18), 
                ("Sessions Remaining", False, 22), ("Hours Remaining", False, 20), ("Notes", False, 30)]
    style_header(ws_pkg, pkg_cols)
    
    pkg_rows = []
    for i in range(1, 11):
        phone = f"+155500000{i:02d}"
        if is_success:
            if i % 2 == 0:
                pkg_rows.append([phone, "10-Session Pro Coaching", "coaching", 5, "", f"Row {i} package"])
            else:
                pkg_rows.append([phone, "20-Hour Winter Pass", "simulator", "", 10.5, f"Row {i} package"])
        else:
            if i == 1:
                pkg_rows.append(["+19999999999", "10-Session Pro Coaching", "coaching", 5, "", "Phone not in Customers"]) # Error: Customer not found
            elif i == 2:
                pkg_rows.append([phone, "Non-existent Package", "coaching", 5, "", "Package doesn't exist"]) # Error: Package not found
            elif i == 3:
                pkg_rows.append([phone, "10-Session Pro Coaching", "invalid_type", 5, "", "Bad type"]) # Error: Invalid type
            elif i == 4:
                pkg_rows.append([phone, "10-Session Pro Coaching", "coaching", "five", "", "Bad number"]) # Error: Sessions not a number
            else:
                pkg_rows.append([phone, "10-Session Pro Coaching", "coaching", 5, "", "Good row"])

    add_rows(ws_pkg, pkg_rows)

    # -------------------
    # Bookings
    # -------------------
    ws_book = wb.create_sheet("Bookings")
    book_cols = [("Customer Phone", True, 18), ("Booking Type", True, 18), ("Category Name", False, 24), 
                 ("Start Time", True, 24), ("End Time", True, 24), ("Status", False, 18), ("Total Price", False, 16)]
    style_header(ws_book, book_cols)
    
    book_rows = []
    for i in range(1, 11):
        phone = f"+155500000{i:02d}"
        start = f"2024-10-{i:02d} 10:00"
        end = f"2024-10-{i:02d} 11:00"
        if is_success:
            if i % 3 == 0:
                book_rows.append([phone, "simulator", "", start, end, "completed", "45.00"])
            elif i % 3 == 1:
                book_rows.append([phone, "coaching", "", start, end, "confirmed", "100.00"])
            else:
                book_rows.append([phone, "category", "Table Tennis", start, end, "no_show", "15.00"])
        else:
            if i == 1:
                book_rows.append([phone, "simulator", "", end, start, "completed", "0.00"]) # Error: End before start
            elif i == 2:
                book_rows.append([phone, "category", "", start, end, "completed", "0.00"]) # Error: Missing category name
            elif i == 3:
                book_rows.append([phone, "invalid_type", "", start, end, "completed", "0.00"]) # Error: Invalid booking type
            elif i == 4:
                book_rows.append([phone, "simulator", "", "bad-date", end, "completed", "0.00"]) # Error: Invalid date format
            else:
                book_rows.append([phone, "simulator", "", start, end, "completed", "45.00"])

    add_rows(ws_book, book_rows)

    # -------------------
    # Events
    # -------------------
    ws_evt = wb.create_sheet("Events")
    evt_cols = [("Customer Phone", True, 18), ("Event Name", True, 30), ("Occurrence Date", True, 20), ("Registration Status", False, 22)]
    style_header(ws_evt, evt_cols)
    
    evt_rows = []
    today_str = timezone.now().date().strftime("%Y-%m-%d")
    for i in range(1, 11):
        phone = f"+155500000{i:02d}"
        if is_success:
            status = "registered" if i % 2 == 0 else "showed_up"
            evt_rows.append([phone, "Saturday Morning Clinic", today_str, status])
        else:
            if i == 1:
                evt_rows.append([phone, "Non-existent Event", today_str, "registered"]) # Error: Event not found
            elif i == 2:
                evt_rows.append(["+19999999999", "Saturday Morning Clinic", today_str, "registered"]) # Error: Customer not found
            elif i == 3:
                evt_rows.append([phone, "", today_str, "registered"]) # Error: Missing event name
            elif i == 4:
                evt_rows.append([phone, "Saturday Morning Clinic", "bad-date", "invalid_status"]) # Error: Invalid date / ignored status
            else:
                evt_rows.append([phone, "Saturday Morning Clinic", today_str, "registered"])

    add_rows(ws_evt, evt_rows)

    wb.save(filename)
    print(f"Created {filename}")

if __name__ == "__main__":
    setup_dummy_data()
    create_excel("success_upload.xlsx", is_success=True)
    create_excel("fail_upload.xlsx", is_success=False)
